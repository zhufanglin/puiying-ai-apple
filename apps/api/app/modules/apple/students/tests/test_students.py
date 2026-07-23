from __future__ import annotations

import os
import shutil
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.modules.apple.students.file_store import StudentFileStore
from app.modules.apple.students.attendance_service import AttendanceService
from app.modules.apple.students.certificate_service import CertificateService
from app.modules.apple.students.models import Attendance, CertificateRequest, Student
from app.modules.apple.students.photo_service import StudentPhotoService
from app.modules.apple.students.repository import StudentRepository
from app.modules.apple.students.schemas import CertificateCreate, StudentCreate, StudentUpdate
from app.modules.apple.students.score_service import ScoreService
from app.modules.apple.students.student_service import StudentConflictError, StudentService


class StudentServicesTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        source_data = Path(__file__).resolve().parents[5] / "data"
        data_dir = Path(self.temp.name) / "data"
        data_dir.mkdir()
        shutil.copy2(source_data / "apple_students_state.json", data_dir / "apple_students_state.json")
        self.previous_data_dir = os.environ.get("APPLE_DATA_DIR")
        os.environ["APPLE_DATA_DIR"] = str(data_dir)
        self.store = StudentFileStore()
        self.repository = StudentRepository(self.store)
        self.students = StudentService(self.repository)
        self.attendance = AttendanceService(self.repository)
        self.certificates = CertificateService(self.repository)
        self.photos = StudentPhotoService(self.repository)
        self.scores = ScoreService(self.repository)

    def tearDown(self) -> None:
        if self.previous_data_dir is None:
            os.environ.pop("APPLE_DATA_DIR", None)
        else:
            os.environ["APPLE_DATA_DIR"] = self.previous_data_dir
        self.temp.cleanup()

    def test_list_and_summary(self) -> None:
        rows = self.students.list()
        self.assertGreaterEqual(len(rows), 3)
        self.assertEqual(self.students.summary()["activeStudents"], 4)
        self.assertGreaterEqual(self.students.summary()["monthlyAttendanceExceptions"], 1)

    def test_list_parent_phones_by_class(self) -> None:
        rows = self.students.list_parent_phones("中四A班")

        source_rows = [
            row
            for row in self.students.list(class_name="中四A班", status="active")
            if (row.get("parentPhone") or "").strip()
        ]
        self.assertEqual(len(rows), len(source_rows))
        self.assertTrue(rows)
        self.assertEqual(
            rows[0],
            {
                "student_name": source_rows[0]["nameZh"],
                "parent_name": source_rows[0].get("parentName"),
                "phone": source_rows[0]["parentPhone"].strip(),
            },
        )

    def test_parent_phone_route_is_registered(self) -> None:
        from app.modules.apple.students.router import router

        paths = {route.path for route in router.routes}
        self.assertIn("/{class_name}/parent-phones", paths)

    def test_create_update_and_delete(self) -> None:
        created = self.students.create(StudentCreate(studentNo="S26999", nameZh="测试学生", nameEn="Test Student", className="3C", parentEmail="parent@example.com"))
        self.assertEqual(created["studentNo"], "S26999")
        updated = self.students.update(created["id"], StudentUpdate(className="3D", parentPhone="61230000"))
        self.assertEqual(updated["className"], "3D")
        self.assertTrue(self.students.delete(created["id"])["deleted"])
        self.assertFalse(any(row["id"] == created["id"] for row in self.students.list()))

    def test_duplicate_student_number_is_rejected(self) -> None:
        with self.assertRaises(StudentConflictError):
            self.students.create(StudentCreate(studentNo="STU-2024001", nameZh="重复", className="中四A班"))

    def test_student_excel_import_upserts_and_reports_errors(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["学号", "姓名", "班级", "入学日期", "家长电邮"])
        sheet.append(["STU-2024001", "陳小明", "中四A班", "2025-09-01", "parent.stu2024001@example.edu.hk"])
        sheet.append(["STU-2024998", "新同学", "中一A班", "2026-09-01", "new.parent@example.edu.hk"])
        sheet.append(["", "缺学号", "1A", "2026-09-01", "bad@example.edu.hk"])
        stream = BytesIO(); workbook.save(stream)
        result = self.students.import_excel(stream.getvalue(), "students.xlsx")
        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["updated"], 1)
        self.assertEqual(len(result["errors"]), 1)
        self.assertEqual(self.students.list(search="STU-2024998")[0]["nameZh"], "新同学")

    def test_attendance_excel_import_and_anomaly_detection(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["日期", "状态", "备注", "学号"])
        sheet.append(["2026-07-10", "出勤", "", "STU-2024001"])
        sheet.append(["2026-07-11", "迟到", "交通延误", "STU-2024001"])
        sheet.append(["2026-07-12", "缺席", "错误学生", "STU-2024002"])
        stream = BytesIO(); workbook.save(stream)
        result = self.attendance.import_excel("stu-0000000001", stream.getvalue(), "attendance.xlsx")
        self.assertEqual(result["imported"], 2)
        self.assertEqual(len(result["errors"]), 1)
        self.assertTrue(any(row["date"] == "2026-07-11" for row in self.attendance.anomalies("stu-0000000001")))

    def test_bulk_attendance_import_across_students(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["学号", "日期", "状态", "备注"])
        sheet.append(["STU-2024001", "2026-08-20", "迟到", "交通延误"])
        sheet.append(["STU-2024002", "2026-08-20", "病假", "已交证明"])
        sheet.append(["STU-9999999", "2026-08-20", "缺席", "不存在的学生"])
        stream = BytesIO(); workbook.save(stream)
        result = self.attendance.import_bulk_excel(stream.getvalue(), "attendance-bulk.xlsx")
        self.assertEqual(result["imported"], 2)
        self.assertEqual(result["affectedStudents"], 2)
        self.assertEqual(len(result["errors"]), 1)

    def test_work_items_can_filter_by_type_class_and_time(self) -> None:
        rows = self.students.work_items(
            category="attendance",
            class_name="中四A班",
            status="absent",
            date_from="2026-07-01",
            date_to="2026-07-31",
        )
        self.assertTrue(rows)
        self.assertTrue(all(row["className"] == "中四A班" and row["status"] == "absent" for row in rows))
        pending = self.students.work_items(category="enrollment_certificate")
        self.assertTrue(all(row["status"] != "generated" for row in pending))

    def test_score_export_filters_by_subject(self) -> None:
        state = self.repository.state()
        state["scoreRecords"].append({
            "id": "score-test-001",
            "studentId": "stu-0000000001",
            "schoolYear": "2025/26",
            "term": "上学期",
            "subject": "中国语文",
            "score": 88,
            "grade": "A",
        })
        self.repository.save(state)

        content, filename = self.scores.export_excel("stu-0000000001", subject="中国语文")
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook["成绩记录"]
        self.assertIn("STU-2024001", filename)
        self.assertEqual(sheet["C5"].value, "中国语文")
        self.assertEqual(sheet.max_row, 5)

    def test_student_photo_upload_and_lookup(self) -> None:
        content = b"\x89PNG\r\n\x1a\n" + b"\x00" * 32
        saved = self.photos.save(content, "student.png", "image/png")
        path, media_type = self.photos.get(saved["photoId"])
        self.assertTrue(path.is_file())
        self.assertEqual(media_type, "image/png")
        self.assertTrue(saved["photoUrl"].startswith("/api/v1/apple/students/photos/"))

    def test_certificate_request_and_pdf_generation(self) -> None:
        request = self.certificates.create("stu-0000000001", CertificateCreate(certificateType="enrollment_bilingual", language="bilingual", purpose="测试用途"))
        pdf_path, updated = self.certificates.generate_pdf("stu-0000000001", request["id"])
        self.assertTrue(pdf_path.is_file())
        self.assertGreater(pdf_path.stat().st_size, 500)
        self.assertEqual(updated["status"], "generated")

    def test_orm_table_contract(self) -> None:
        self.assertEqual(Student.__tablename__, "apple_students")
        self.assertEqual(Attendance.__tablename__, "apple_attendance")
        self.assertEqual(CertificateRequest.__tablename__, "apple_certificate_requests")
        self.assertIn("student_no", Student.__table__.columns)


if __name__ == "__main__":
    unittest.main()
