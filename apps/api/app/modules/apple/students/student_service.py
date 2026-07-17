from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.modules.apple.students.file_store import now_iso
from app.modules.apple.students.repository import StudentRepository
from app.modules.apple.students.schemas import StudentCreate, StudentUpdate


class StudentNotFoundError(LookupError):
    pass


class StudentConflictError(ValueError):
    pass


HEADER_ALIASES = {
    "学号": "studentNo", "學生編號": "studentNo", "student_no": "studentNo", "studentno": "studentNo",
    "姓名": "nameZh", "中文姓名": "nameZh", "name": "nameZh", "name_zh": "nameZh",
    "英文姓名": "nameEn", "name_en": "nameEn",
    "班级": "className", "班級": "className", "class": "className", "class_name": "className",
    "状态": "status", "狀態": "status", "status": "status",
    "入学日期": "admissionDate", "入學日期": "admissionDate", "admission_date": "admissionDate",
    "家长姓名": "parentName", "家長姓名": "parentName", "parent_name": "parentName",
    "家长电话": "parentPhone", "家長電話": "parentPhone", "parent_phone": "parentPhone",
    "家长电邮": "parentEmail", "家長電郵": "parentEmail", "parent_email": "parentEmail",
    "照片地址": "photoUrl", "照片網址": "photoUrl", "photo_url": "photoUrl", "photourl": "photoUrl",
}

STATUS_ALIASES = {
    "active": "active", "在读": "active", "在讀": "active",
    "suspended": "suspended", "停学": "suspended", "停學": "suspended",
    "withdrawn": "withdrawn", "离校": "withdrawn", "離校": "withdrawn",
}


class StudentService:
    def __init__(self, repository: StudentRepository | None = None) -> None:
        self.repository = repository or StudentRepository()

    def list(self, *, search: str | None = None, class_name: str | None = None, status: str | None = None) -> list[dict[str, Any]]:
        state = self.repository.state()
        return self.repository.list_students(state, search=search, class_name=class_name, status=status)

    def get(self, student_id: str) -> dict[str, Any]:
        state = self.repository.state()
        student = self.repository.get_student(state, student_id)
        if not student:
            raise StudentNotFoundError(student_id)
        return {
            **student,
            "attendance": self.repository.attendance(state, student_id),
            "certificates": self.repository.certificates(state, student_id),
            "scores": self.repository.scores(state, student_id),
        }

    def create(self, body: StudentCreate) -> dict[str, Any]:
        state = self.repository.state()
        if self.repository.get_by_no(state, body.studentNo):
            raise StudentConflictError("学号已经存在")
        row = self.repository.create_student(state, body.model_dump(mode="json"))
        self.repository.storage.audit(state, "student.created", "student", row["id"], {"studentNo": row["studentNo"]})
        self.repository.save(state)
        return row

    def update(self, student_id: str, body: StudentUpdate) -> dict[str, Any]:
        state = self.repository.state()
        row = self.repository.get_student(state, student_id)
        if not row:
            raise StudentNotFoundError(student_id)
        values = body.model_dump(mode="json", exclude_unset=True)
        updated = self.repository.update_student(row, values)
        self.repository.storage.audit(state, "student.updated", "student", student_id, {"fields": sorted(values)})
        self.repository.save(state)
        return updated

    def delete(self, student_id: str) -> dict[str, Any]:
        state = self.repository.state()
        row = self.repository.get_student(state, student_id)
        if not row:
            raise StudentNotFoundError(student_id)
        row.update({"status": "deleted", "updatedAt": now_iso(), "updatedBy": "Apple"})
        self.repository.storage.audit(state, "student.deleted", "student", student_id, {"studentNo": row["studentNo"]})
        self.repository.save(state)
        return {"id": student_id, "deleted": True}

    def summary(self) -> dict[str, int]:
        state = self.repository.state()
        active = [row for row in state["students"] if row.get("status") == "active"]
        current_month = now_iso()[:7]
        abnormal = sum(
            row["date"].startswith(current_month) and row["status"] != "present"
            for row in state["attendanceRecords"]
        )
        pending_reissue = sum(
            row["certificateType"] == "transcript_reissue" and row["status"] != "generated"
            for row in state["certificateRequests"]
        )
        pending_certificate = sum(
            row["certificateType"] != "transcript_reissue" and row["status"] != "generated"
            for row in state["certificateRequests"]
        )
        return {
            "activeStudents": len(active),
            "monthlyAttendanceExceptions": abnormal,
            "pendingTranscriptReissues": pending_reissue,
            "pendingEnrollmentCertificates": pending_certificate,
        }

    def work_items(
        self,
        *,
        category: str,
        search: str | None = None,
        class_name: str | None = None,
        status: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> list[dict[str, Any]]:
        if category not in {"attendance", "transcript_reissue", "enrollment_certificate"}:
            raise ValueError("不支持的待办分类")
        state = self.repository.state()
        students = {
            row["id"]: row
            for row in state["students"]
            if row.get("status") != "deleted"
        }
        items: list[dict[str, Any]] = []
        if category == "attendance":
            for row in state["attendanceRecords"]:
                if row.get("status") == "present" or row.get("studentId") not in students:
                    continue
                student = students[row["studentId"]]
                items.append({
                    "id": row["id"],
                    "category": category,
                    "studentId": student["id"],
                    "studentNo": student["studentNo"],
                    "studentName": student["nameZh"],
                    "className": student["className"],
                    "date": row["date"],
                    "status": row["status"],
                    "detail": row.get("remarks") or "未填写备注",
                })
        else:
            expected_type = "transcript_reissue" if category == "transcript_reissue" else None
            for row in state["certificateRequests"]:
                is_expected = row.get("certificateType") == expected_type if expected_type else row.get("certificateType") != "transcript_reissue"
                if not is_expected or row.get("status") == "generated" or row.get("studentId") not in students:
                    continue
                student = students[row["studentId"]]
                items.append({
                    "id": row["id"],
                    "category": category,
                    "studentId": student["id"],
                    "studentNo": student["studentNo"],
                    "studentName": student["nameZh"],
                    "className": student["className"],
                    "date": row["requestDate"],
                    "status": row["status"],
                    "detail": row.get("purpose") or "未填写用途",
                })

        if search:
            needle = search.casefold()
            items = [
                row for row in items
                if needle in row["studentNo"].casefold()
                or needle in row["studentName"].casefold()
                or needle in row["detail"].casefold()
            ]
        if class_name:
            items = [row for row in items if row["className"] == class_name]
        if status:
            items = [row for row in items if row["status"] == status]
        if date_from:
            items = [row for row in items if row["date"] >= date_from]
        if date_to:
            items = [row for row in items if row["date"] <= date_to]
        return sorted(items, key=lambda row: (row["date"], row["studentNo"]), reverse=True)

    def import_excel(self, content: bytes, filename: str) -> dict[str, Any]:
        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError("无法读取 Excel 文件") from exc
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if not raw_headers:
            raise ValueError("Excel 没有表头")
        headers = [HEADER_ALIASES.get(str(value).strip().casefold(), HEADER_ALIASES.get(str(value).strip(), "")) if value is not None else "" for value in raw_headers]
        if "studentNo" not in headers or "nameZh" not in headers or "className" not in headers:
            raise ValueError("Excel 必须包含学号、姓名和班级")

        state = self.repository.state()
        source_id = self.repository.storage.new_id("file")
        destination = self.repository.storage.upload_dir / f"{source_id}_{Path(filename).name}"
        destination.write_bytes(content)
        state["files"].append({
            "id": source_id, "fileName": Path(filename).name, "fileType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "path": str(destination), "hash": self.repository.storage.hash_file(destination), "status": "uploaded", "createdAt": now_iso(),
        })
        imported = updated = skipped = 0
        errors: list[dict[str, Any]] = []
        for row_number, raw_values in enumerate(rows, start=2):
            values = {headers[index]: raw_values[index] for index in range(min(len(headers), len(raw_values))) if headers[index] and raw_values[index] not in {None, ""}}
            if not values:
                skipped += 1
                continue
            values["studentNo"] = str(values.get("studentNo", "")).strip().upper()
            values["nameZh"] = str(values.get("nameZh", "")).strip()
            values["className"] = str(values.get("className", "")).strip().upper()
            raw_status = str(values.get("status", "active")).strip()
            values["status"] = STATUS_ALIASES.get(raw_status, STATUS_ALIASES.get(raw_status.lower(), raw_status.lower()))
            if hasattr(values.get("admissionDate"), "isoformat"):
                values["admissionDate"] = values["admissionDate"].date().isoformat() if hasattr(values["admissionDate"], "date") else values["admissionDate"].isoformat()
            values["sourceFileId"] = source_id
            try:
                validated = StudentCreate.model_validate(values).model_dump(mode="json")
            except Exception as exc:
                errors.append({"row": row_number, "message": str(exc).splitlines()[0]})
                continue
            existing = self.repository.get_by_no(state, validated["studentNo"])
            if existing:
                self.repository.update_student(existing, {**validated, "sourceFileId": source_id})
                updated += 1
            else:
                self.repository.create_student(state, {**validated, "sourceFileId": source_id})
                imported += 1
        self.repository.storage.audit(state, "students.excel_imported", "file", source_id, {"imported": imported, "updated": updated, "skipped": skipped, "errors": len(errors)})
        self.repository.save(state)
        return {"sourceFileId": source_id, "imported": imported, "updated": updated, "skipped": skipped, "errors": errors}
