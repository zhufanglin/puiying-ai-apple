from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.modules.apple.file_store import now_iso
from app.modules.apple.students.repository import StudentRepository
from app.modules.apple.students.schemas import AttendanceStatus
from app.modules.apple.students.student_service import StudentNotFoundError


ATTENDANCE_LABELS = {
    "present": "present", "出勤": "present",
    "late": "late", "迟到": "late", "遲到": "late",
    "absent": "absent", "缺席": "absent",
    "sick_leave": "sick_leave", "病假": "sick_leave",
}


class AttendanceService:
    def __init__(self, repository: StudentRepository | None = None) -> None:
        self.repository = repository or StudentRepository()

    def list(self, student_id: str) -> list[dict[str, Any]]:
        state = self.repository.state()
        if not self.repository.get_student(state, student_id):
            raise StudentNotFoundError(student_id)
        return self.repository.attendance(state, student_id)

    def import_excel(self, student_id: str, content: bytes, filename: str) -> dict[str, Any]:
        state = self.repository.state()
        student = self.repository.get_student(state, student_id)
        if not student:
            raise StudentNotFoundError(student_id)
        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError("无法读取考勤 Excel 文件") from exc
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if not raw_headers:
            raise ValueError("Excel 没有表头")
        aliases = {"日期": "date", "date": "date", "状态": "status", "狀態": "status", "status": "status", "备注": "remarks", "備註": "remarks", "remarks": "remarks", "学号": "studentNo", "學號": "studentNo", "student_no": "studentNo"}
        headers = [aliases.get(str(value).strip().casefold(), aliases.get(str(value).strip(), "")) if value is not None else "" for value in raw_headers]
        if "date" not in headers or "status" not in headers:
            raise ValueError("考勤 Excel 必须包含日期和状态")
        source_id = self.repository.storage.new_id("file")
        destination = self.repository.storage.upload_dir / f"{source_id}_{Path(filename).name}"
        destination.write_bytes(content)
        state["files"].append({"id": source_id, "fileName": Path(filename).name, "fileType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "path": str(destination), "hash": self.repository.storage.hash_file(destination), "status": "uploaded", "createdAt": now_iso()})
        imported = updated = skipped = 0
        errors: list[dict[str, Any]] = []
        for row_number, raw_values in enumerate(rows, start=2):
            values = {headers[index]: raw_values[index] for index in range(min(len(headers), len(raw_values))) if headers[index] and raw_values[index] not in {None, ""}}
            if not values:
                skipped += 1
                continue
            if values.get("studentNo") and str(values["studentNo"]).strip().upper() != student["studentNo"]:
                errors.append({"row": row_number, "message": "学号与当前学生不符"})
                continue
            date_value = values.get("date")
            if hasattr(date_value, "isoformat"):
                date_value = date_value.date().isoformat() if hasattr(date_value, "date") else date_value.isoformat()
            status = ATTENDANCE_LABELS.get(str(values.get("status", "")).strip())
            if not date_value or status not in set(AttendanceStatus):
                errors.append({"row": row_number, "message": "日期或考勤状态无效"})
                continue
            _, was_updated = self.repository.upsert_attendance(state, {"studentId": student_id, "date": str(date_value), "status": status, "remarks": str(values.get("remarks", "")), "sourceFileId": source_id, "reviewStatus": "confirmed"})
            if was_updated: updated += 1
            else: imported += 1
        self.repository.storage.audit(state, "attendance.excel_imported", "student", student_id, {"studentId": student_id, "sourceFileId": source_id, "imported": imported, "updated": updated, "errors": len(errors)})
        self.repository.save(state)
        return {"sourceFileId": source_id, "imported": imported, "updated": updated, "skipped": skipped, "errors": errors}

    def import_bulk_excel(self, content: bytes, filename: str) -> dict[str, Any]:
        """按学号跨学生批量导入考勤，供学生总览页面使用。"""
        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError("无法读取考勤 Excel 文件") from exc
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        aliases = {
            "学号": "studentNo", "學號": "studentNo", "student_no": "studentNo", "studentno": "studentNo",
            "日期": "date", "date": "date",
            "状态": "status", "狀態": "status", "status": "status",
            "备注": "remarks", "備註": "remarks", "remarks": "remarks",
        }
        headers: list[str] = []
        header_row_number = 0
        for header_row_number, raw_headers in enumerate(rows, start=1):
            headers = [
                aliases.get(str(value).strip().casefold(), aliases.get(str(value).strip(), ""))
                if value is not None else ""
                for value in raw_headers
            ]
            if {"studentNo", "date", "status"}.issubset(headers):
                break
            if header_row_number >= 10:
                headers = []
                break
        if not headers or not {"studentNo", "date", "status"}.issubset(headers):
            raise ValueError("批量考勤 Excel 前 10 行必须包含学号、日期和状态表头")

        state = self.repository.state()
        source_id = self.repository.storage.new_id("file")
        destination = self.repository.storage.upload_dir / f"{source_id}_{Path(filename).name}"
        destination.write_bytes(content)
        state["files"].append({
            "id": source_id,
            "fileName": Path(filename).name,
            "fileType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "path": str(destination),
            "hash": self.repository.storage.hash_file(destination),
            "status": "uploaded",
            "createdAt": now_iso(),
        })

        imported = updated = skipped = 0
        errors: list[dict[str, Any]] = []
        affected_students: set[str] = set()
        for row_number, raw_values in enumerate(rows, start=header_row_number + 1):
            values = {
                headers[index]: raw_values[index]
                for index in range(min(len(headers), len(raw_values)))
                if headers[index] and raw_values[index] not in {None, ""}
            }
            if not values:
                skipped += 1
                continue
            student_no = str(values.get("studentNo", "")).strip().upper()
            student = self.repository.get_by_no(state, student_no)
            if not student:
                errors.append({"row": row_number, "message": f"找不到学号 {student_no or '（空白）'}"})
                continue
            date_value = values.get("date")
            if hasattr(date_value, "isoformat"):
                date_value = date_value.date().isoformat() if hasattr(date_value, "date") else date_value.isoformat()
            raw_status = str(values.get("status", "")).strip()
            attendance_status = ATTENDANCE_LABELS.get(raw_status, ATTENDANCE_LABELS.get(raw_status.lower()))
            if not date_value or attendance_status not in set(AttendanceStatus):
                errors.append({"row": row_number, "message": "日期或考勤状态无效"})
                continue
            _, was_updated = self.repository.upsert_attendance(state, {
                "studentId": student["id"],
                "date": str(date_value),
                "status": attendance_status,
                "remarks": str(values.get("remarks", "")),
                "sourceFileId": source_id,
                "reviewStatus": "confirmed",
            })
            affected_students.add(student["id"])
            if was_updated:
                updated += 1
            else:
                imported += 1

        self.repository.storage.audit(state, "attendance.bulk_excel_imported", "file", source_id, {
            "sourceFileId": source_id,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": len(errors),
            "affectedStudents": len(affected_students),
        })
        self.repository.save(state)
        return {
            "sourceFileId": source_id,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
            "affectedStudents": len(affected_students),
        }

    def anomalies(self, student_id: str) -> list[dict[str, Any]]:
        return [row for row in self.list(student_id) if row["status"] != "present"]
