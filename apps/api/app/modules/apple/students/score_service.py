"""學生既有成績匯出，以及第二組共用的成績匯入/統計算法。"""
from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from statistics import mean
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.modules.apple.students.repository import StudentRepository
from app.modules.apple.students.student_service import StudentNotFoundError


@dataclass(frozen=True, slots=True)
class ParsedScore:
    row: int
    student_no: str
    student_name: str | None
    subject: str
    score: Decimal
    full_mark: Decimal


_HEADER_ALIASES = {
    "學號": "student_no",
    "学号": "student_no",
    "student_no": "student_no",
    "student no": "student_no",
    "姓名": "student_name",
    "學生姓名": "student_name",
    "学生姓名": "student_name",
    "name": "student_name",
    "科目": "subject",
    "subject": "subject",
    "成績": "score",
    "成绩": "score",
    "分數": "score",
    "分数": "score",
    "score": "score",
    "滿分": "full_mark",
    "满分": "full_mark",
    "full_mark": "full_mark",
    "full mark": "full_mark",
}
_SUBJECT_FULL_MARK = re.compile(r"^\s*(?P<subject>.+?)\s*[（(]\s*(?P<mark>\d+(?:\.\d+)?)\s*[)）]\s*$")


def _normalise_header(value: Any) -> str:
    text = str(value or "").strip()
    return _HEADER_ALIASES.get(text.casefold(), _HEADER_ALIASES.get(text, text))


def _subject_header(value: Any) -> tuple[str, Decimal]:
    text = str(value or "").strip()
    match = _SUBJECT_FULL_MARK.match(text)
    if not match:
        return text, Decimal("100")
    return match.group("subject").strip(), Decimal(match.group("mark"))


def _decimal_score(value: Any) -> Decimal:
    if isinstance(value, bool):
        raise ValueError("分數不可為布林值")
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError("分數必須是數字") from exc
    if not number.is_finite():
        raise ValueError("分數必須是有限數字")
    return number


def _validated_score(value: Any, full_mark_value: Any) -> tuple[Decimal, Decimal]:
    score = _decimal_score(value)
    full_mark = _decimal_score(full_mark_value)
    if full_mark <= 0:
        raise ValueError("滿分必須大於 0")
    if score < 0 or score > full_mark:
        raise ValueError(f"分數須介乎 0 至 {full_mark:g}")
    return score, full_mark


def parse_score_workbook(content: bytes) -> tuple[list[ParsedScore], list[dict[str, Any]]]:
    """解析長表或寬表 Excel；壞資料逐格回報，正確資料仍可匯入。"""

    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("無法讀取成績 Excel 文件") from exc

    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    raw_headers = next(iterator, None)
    if not raw_headers:
        raise ValueError("Excel 沒有表頭")

    headers = [_normalise_header(value) for value in raw_headers]
    if "student_no" not in headers:
        raise ValueError("成績 Excel 必須包含學號")

    long_form = "subject" in headers and "score" in headers
    if not long_form:
        subject_indexes = [
            index
            for index, header in enumerate(headers)
            if header and header not in {"student_no", "student_name", "full_mark"}
        ]
        if not subject_indexes:
            raise ValueError("成績 Excel 必須包含科目及分數")

    parsed: list[ParsedScore] = []
    errors: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row_number, row in enumerate(iterator, start=2):
        if not row or all(value in {None, ""} for value in row):
            continue
        values = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
        student_no = str(values.get("student_no") or "").strip().upper()
        student_name = str(values.get("student_name") or "").strip() or None
        if not student_no:
            errors.append({"row": row_number, "message": "缺少學號"})
            continue

        candidates: list[tuple[str, Any, Any]] = []
        if long_form:
            subject = str(values.get("subject") or "").strip()
            if not subject:
                errors.append({"row": row_number, "message": "缺少科目"})
                continue
            candidates.append((subject, values.get("score"), values.get("full_mark") or 100))
        else:
            for index in subject_indexes:
                raw_score = row[index] if index < len(row) else None
                if raw_score in {None, ""}:
                    continue
                subject, full_mark = _subject_header(raw_headers[index])
                candidates.append((subject, raw_score, full_mark))

        if not candidates:
            errors.append({"row": row_number, "message": "沒有可匯入的分數"})
            continue

        for subject, raw_score, raw_full_mark in candidates:
            try:
                score, full_mark = _validated_score(raw_score, raw_full_mark)
            except ValueError as exc:
                errors.append({"row": row_number, "subject": subject, "message": str(exc)})
                continue
            key = (student_no.casefold(), subject.casefold())
            if key in seen:
                errors.append({"row": row_number, "subject": subject, "message": "檔案內學號及科目重複"})
                continue
            seen.add(key)
            parsed.append(ParsedScore(row_number, student_no, student_name, subject, score, full_mark))

    if not parsed and not errors:
        raise ValueError("Excel 沒有成績資料")
    return parsed, errors


def _as_percentage(row: dict[str, Any]) -> float:
    full_mark = float(row.get("full_mark", row.get("fullMark", 100)) or 100)
    return round(float(row.get("score", 0)) / full_mark * 100, 4) if full_mark > 0 else 0.0


def _band(percentage: float) -> str:
    if percentage >= 85:
        return "A"
    if percentage >= 70:
        return "B"
    if percentage >= 60:
        return "C"
    return "D"


def _student_percentages(records: Iterable[dict[str, Any]]) -> dict[str, float]:
    grouped: dict[str, list[float]] = {}
    for row in records:
        student_id = str(row.get("student_id", row.get("studentId", "")))
        if student_id:
            grouped.setdefault(student_id, []).append(_as_percentage(row))
    return {student_id: mean(values) for student_id, values in grouped.items() if values}


def calc_class_stats(records: Iterable[dict[str, Any]]) -> dict[str, Any]:
    """以每名學生的科目百分比平均值計算班級統計。"""

    rows = list(records)
    percentages = _student_percentages(rows)
    values = list(percentages.values())
    bands = {"A": 0, "B": 0, "C": 0, "D": 0}
    for value in values:
        bands[_band(value)] += 1

    subject_groups: dict[str, list[tuple[float, float]]] = {}
    for row in rows:
        subject = str(row.get("subject", ""))
        if subject:
            subject_groups.setdefault(subject, []).append((float(row.get("score", 0)), _as_percentage(row)))

    subject_averages = [
        {
            "subject": subject,
            "average_score": round(mean(score for score, _ in items), 2),
            "average_percentage": round(mean(percent for _, percent in items), 2),
        }
        for subject, items in sorted(subject_groups.items())
    ]
    rankings = sorted(percentages.items(), key=lambda item: (-item[1], item[0]))
    return {
        "student_count": len(values),
        "average": round(mean(values), 2) if values else 0.0,
        "highest": round(max(values), 2) if values else 0.0,
        "lowest": round(min(values), 2) if values else 0.0,
        "pass_rate": round(sum(value >= 60 for value in values) / len(values) * 100, 2) if values else 0.0,
        "bands": bands,
        "subject_averages": subject_averages,
        "rankings": [
            {"student_id": student_id, "rank": index, "average_percentage": round(value, 2)}
            for index, (student_id, value) in enumerate(rankings, start=1)
        ],
    }


def calc_student_stats(records: Iterable[dict[str, Any]], student_id: str) -> dict[str, Any] | None:
    """計算個人總分、班級排名及強弱科目。records 應包含同班全體資料。"""

    rows = list(records)
    target = [row for row in rows if str(row.get("student_id", row.get("studentId", ""))) == student_id]
    if not target:
        return None

    class_stats = calc_class_stats(rows)
    rank_row = next((row for row in class_stats["rankings"] if row["student_id"] == student_id), None)
    subjects = [
        {
            "subject": str(row.get("subject", "")),
            "score": float(row.get("score", 0)),
            "full_mark": float(row.get("full_mark", row.get("fullMark", 100))),
            "percentage": round(_as_percentage(row), 2),
        }
        for row in target
    ]
    subjects.sort(key=lambda item: item["subject"])
    strongest = max(subjects, key=lambda item: item["percentage"])
    weakest = min(subjects, key=lambda item: item["percentage"])
    return {
        "student_id": student_id,
        "total_score": round(sum(item["score"] for item in subjects), 2),
        "total_full_mark": round(sum(item["full_mark"] for item in subjects), 2),
        "average_percentage": rank_row["average_percentage"] if rank_row else 0.0,
        "rank": rank_row["rank"] if rank_row else None,
        "class_size": class_stats["student_count"],
        "strongest_subject": strongest["subject"],
        "weakest_subject": weakest["subject"],
        "subjects": subjects,
    }


class ScoreService:
    """保留學生詳情頁既有的檔案型成績查詢/匯出契約。"""

    def __init__(self, repository: StudentRepository | None = None) -> None:
        self.repository = repository or StudentRepository()

    def list(
        self,
        student_id: str,
        *,
        school_year: str | None = None,
        term: str | None = None,
        subject: str | None = None,
        search: str | None = None,
    ) -> list[dict[str, Any]]:
        state = self.repository.state()
        if not self.repository.get_student(state, student_id):
            raise StudentNotFoundError(student_id)
        rows = list(self.repository.scores(state, student_id))
        if school_year:
            rows = [row for row in rows if row.get("schoolYear") == school_year]
        if term:
            rows = [row for row in rows if row.get("term") == term]
        if subject:
            rows = [row for row in rows if row.get("subject") == subject]
        if search:
            needle = search.casefold()
            rows = [
                row for row in rows
                if needle in str(row.get("subject", "")).casefold()
                or needle in str(row.get("grade", "")).casefold()
            ]
        return sorted(rows, key=lambda row: (row.get("schoolYear", ""), row.get("term", ""), row.get("subject", "")), reverse=True)

    def export_excel(
        self,
        student_id: str,
        *,
        school_year: str | None = None,
        term: str | None = None,
        subject: str | None = None,
    ) -> tuple[bytes, str]:
        state = self.repository.state()
        student = self.repository.get_student(state, student_id)
        if not student:
            raise StudentNotFoundError(student_id)
        rows = self.list(student_id, school_year=school_year, term=term, subject=subject)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "成绩记录"
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:E1")
        sheet["A1"] = f"{student['nameZh']}（{student['studentNo']}）成绩记录"
        sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="0F766E")
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.append(["筛选条件", f"学年：{school_year or '全部'}", f"学期：{term or '全部'}", f"科目：{subject or '全部'}", ""])
        sheet.append([])
        sheet.append(["学年", "学期", "科目", "分数", "等级"])
        for row in rows:
            sheet.append([row.get("schoolYear"), row.get("term"), row.get("subject"), row.get("score"), row.get("grade")])
        header = sheet[4]
        for cell in header:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="155E75")
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:E{max(sheet.max_row, 4)}"
        widths = {"A": 16, "B": 14, "C": 22, "D": 12, "E": 12}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
            row[3].number_format = "0.00"
        output = BytesIO()
        workbook.save(output)
        suffix = "_".join(part for part in (school_year, term, subject) if part) or "all"
        filename = f"{student['studentNo']}_scores_{suffix}.xlsx"
        return output.getvalue(), filename
